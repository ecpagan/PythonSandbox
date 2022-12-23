import openpyxl

# write workbook
wb = openpyxl.Workbook()
first_sheet = wb.get_sheet_names()[0]
sheet = wb.get_sheet_by_name(first_sheet)
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'
wb.save('example.xlsx')

sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
sheet2.title = 'My new sheet name'  # will replace the sheet title for that tab
print(wb.get_sheet_names())
wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My other sheet')
wb.save('example3.xlsx')


# read workbook
workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.get_sheet_names())
sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
cell = sheet['A1']
cell_value = cell.value

for i in range(1, 8):
    print(sheet.cell(row=i, column=2).value)

